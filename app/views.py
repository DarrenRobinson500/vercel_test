from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.http import HttpResponse
import openpyxl as xl
from .forms import *
# import requests
from pandas import ExcelWriter

import socket

Event_Tuple = namedtuple('Event_Tuple', ['date', 'description'])
Dog_Diary = namedtuple('Dog_Diary', ['date', 'bookings'])

nav_bar_items = ["notes", "diary", "events", "quotes", "birthdays", "shopping", "wordle", "wordle_remaining", "word", "dogs", "dog_diary"]

def home(request):
    if not request.user.is_authenticated: return redirect("login")
    home_pc = socket.gethostname() == "Mum_and_Dads"
    parent_note = Note.objects.exclude(parent__isnull=False).first()
    children_notes = Note.objects.filter(parent=parent_note).order_by("order")

    context = {"nav_bar_items": nav_bar_items, 'home_pc': home_pc, "children_notes": children_notes}
    return render(request, "home.html", context)

# -----------------------------
# --------AUTHENTICATION=------
# -----------------------------

def login_user(request):
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('home')
        else:
            messages.success(request, ("Error logging in."))
            return redirect('login')
    else:
        context = {}
        return render(request, 'login.html', context)

def logout_user(request):
    logout(request)
    return redirect("login")

# -----------------------------
# --------UTILITIES------------
# -----------------------------

def downloadpage(request):
    context = {}
    return render(request, 'down_load.html', context)

def downloadexcel(request):
    if not request.user.is_authenticated: return redirect("login")
    if not socket.gethostname() == "Mum_and_Dads": return redirect("notes")

    writer = ExcelWriter('my_data.xlsx', engine='xlsxwriter')
    all_models = \
        [Category, Diary, Note, Quote, Birthday, Dog, Booking, Event, TH, Player, Shopping, Shop, Timer, TimerElement,
         New_Tide, Tide_Date, Weather, Wordle, TennisMatch, TennisGame, General]

    print("All models:", all_models)
    for count, model in enumerate(all_models, 1):
        print("Saving:", model)
        data = model.objects.all()
        df = pd.DataFrame(list(data.values()))
        df.to_excel(writer, sheet_name=f'{model.string_name}', index=False)
    writer.close()

    # Create an HttpResponse object with the Excel file
    response = HttpResponse(open('my_data.xlsx', 'rb').read(), content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="my_data.xlsx"'

    return response

def load_data(request):
    file = "excel/data.xlsx"
    wb = xl.load_workbook(file)
    sheet_names = wb.sheetnames

    context = {"sheet_names": sheet_names, }
    return render(request, "load_data.html", context)

def load_data_ind(request, model_name):
    # model_name = "Dog"
    file = "excel/data.xlsx"
    wb = xl.load_workbook(file)
    sheet = wb[model_name]

    model = None
    for model_x in [Category, Diary, Note, Quote, Birthday, Dog, Booking, Event, TH, Player, Shopping, Shop, Timer, TimerElement,
         New_Tide, Tide_Date, Weather, Wordle, TennisMatch, TennisGame, General]:
        if model_x.string_name == model_name:
            model = model_x

    column_headings = []
    for cell in sheet[1]: column_headings.append(cell.value)

    table_data = []
    for row in sheet.iter_rows():
        row_data = []
        for cell in row:
            row_data.append(cell.value)
        current_id = row[0].value
        print("Current id:", current_id)
        existing_records = 1
        if type(current_id) is int:
            existing_records = len(model.objects.filter(id=current_id))
        if existing_records == 0:
            add_record(model, column_headings, row_data)
        row_data.append(existing_records)
        table_data.append(row_data)

    # Loop through again to include the links
    link_pos = None
    if model.string_name == "Note":
        link_pos = column_headings.index("parent_id")
        link_model = Note
        link_name = "parent"
    if model.string_name == "Booking":
        link_pos = column_headings.index("dog_id")
        link_model = Dog
        link_name = "dog"
    if link_pos:
        for row in sheet.iter_rows():
            current_id = row[0].value
            if type(current_id) is int:
                link_id = row[link_pos].value
                if type(link_id) is int:
                    create_link(model, link_model, link_name, current_id, link_id)

    context = {"heading": model_name, "table_data": table_data,}
    return render(request, "table.html", context)

def add_record(model, headings, values):
    new_record = model(id=values[0])
    new_record.save()
    for field_name, value in zip(headings, values):
        if field_name[-3:] == "_id":
            print("Foreign key detected:", field_name)
        elif hasattr(model, field_name):
            setattr(new_record, field_name, value)
            print("Update data:", model, field_name, value)
        else:
            print(f"{field_name} does not exist in {model.string_name}")
    new_record.save()

def create_link(own_model, link_model, field, own_id, link_id, ):
    record = own_model.objects.get(id=own_id)
    link = link_model.objects.get(id=link_id)
    setattr(record, field, link)
    record.save()

# -----------------------------
# --------DOGS----------------
# -----------------------------

def dogs(request):
    if not request.user.is_authenticated: return redirect("login")
    dogs = Dog.objects.all()
    dogs = sorted(dogs, key=lambda d: d.next_booking())
    context = {"dogs": dogs}
    return render(request, 'dogs.html', context)

def dog(request, id):
    if not request.user.is_authenticated: return redirect("login")
    dogs = Dog.objects.all()
    dogs = sorted(dogs, key=lambda d: d.next_booking())
    dog = Dog.objects.get(id=id)
    if request.method == 'POST':
        form = DogForm(request.POST, request.FILES, instance=dog)
        if form.is_valid(): form.save()

    form = DogForm(instance=dog)

    edit_mode = True
    context = {"dog": dog, "dogs": dogs, "form": form, 'edit_mode': edit_mode}
    print("Rendering HTML")
    return render(request, 'dog.html', context)

def dogs_old(request):
    if not request.user.is_authenticated: return redirect("login")
    form = DogForm()
    if request.method == 'POST':
        form = DogForm(request.POST, request.FILES)
        if form.is_valid(): form.save()
    dogs = Dog.objects.all()
    dogs = sorted(dogs, key=lambda d: d.next_booking())
    table_info = []
    for dog in dogs:
        dog_info = [dog.name, dog.id, dog.owners, dog.owners_number, dog.notes, dog.approved, []]
        table_info.append(dog_info)

    today = date.today()
    bookings = Booking.objects.filter(end_date__gte=today).order_by('start_date')
    for booking in bookings:
        for dog_info in table_info:
            print(dog_info[1], booking.dog.id)
            if dog_info[1] == booking.dog.id:
                dog_info[6].append(booking.short_name() + "\n")

    context = {'table_info': table_info, 'title': "Dogs", "form": form, "edit_mode": False, 'people': people()}
    return render(request, 'dog_simple.html', context)

def dog_edit(request, id):
    print("A")
    if not request.user.is_authenticated: return redirect("login")
    dog = Dog.objects.get(id=id)
    if request.method == 'POST':
        print("B")
        form = DogForm(request.POST, request.FILES, instance=dog)
        if form.is_valid(): form.save()
        print("C")
        return redirect("dogs")

    form = DogForm(instance=dog)
    objects = Dog.objects.all()
    objects = sorted(objects, key=lambda d: d.next_booking())

    count = len(objects)
    options = ["Yes", "No", "Limited"]
    context = {'objects': objects, 'title': "Dogs", 'count': count, "dog": dog, "edit_mode": True,  'people': people(),
               'options':options, 'form':form}
    return render(request, 'dog.html', context)

def dog_diary_old(request):
    if not request.user.is_authenticated: return redirect("login")
    general = General.objects.get(name="main")
    today = date.today()

    bookings = Booking.objects.filter(end_date__gte=today).order_by('start_date')
    print("Dog Diary Bookings:", bookings)

    dog_diary = []
    for x in range(general.dog_diary_days):
        day = today + timedelta(days=x)
        day_bookings = bookings.filter(start_date__lte=day).filter(end_date__gte=day)
        print("Dog Diary View (Day):", day, day_bookings)
        new = Dog_Diary(day, day_bookings)
        dog_diary.append(new)
    general = General.objects.get(name="main")
    durations = [30, 60, 90, 120, 150, 180, 360]

    context = {'dog_diary': dog_diary, 'durations': durations, 'general': general}
    return render(request, 'dog_diary.html', context)

def dog_diary(request):
    if not request.user.is_authenticated: return redirect("login")
    general = General.objects.get(name="main")
    today = date.today()

    dog_diary = []
    for x in range(general.dog_diary_days):
        day = today + timedelta(days=x)
        dog_diary.append([day, []])

    bookings = Booking.objects.filter(end_date__gte=today).order_by('start_date')

    for booking in bookings:
        for dog_day in dog_diary:
            # print(dog_day[0], booking.start_date)
            if booking.start_date <= dog_day[0] <= booking.end_date:
                dog_day[1].append(booking.dog.name)

    durations = [30, 60, 90, 120, 150, 180, 360]
    context = {'dog_diary': dog_diary, 'durations': durations, 'general': general}
    return render(request, 'dog_diary.html', context)

def dog_duration(request, dur):
    general = General.objects.get(name="main")
    general.dog_diary_days = dur
    general.save()
    return redirect('dog_diary')

def booking(request, id):
    if not request.user.is_authenticated: return redirect("login")
    dog = Dog.objects.filter(id=id).first()
    if request.method == 'POST':
        form = BookingForm(request.POST or None)
        form.instance.dog = dog
        if form.is_valid():
            new_booking = form.save()
            return redirect("dogs")
        else:
            context = {"dog": dog, "title": f"Booking for {dog.name}", "form": form}
            return render(request, 'booking.html', context)

    context = {"dog": dog, "title": f"Booking for {dog.name}"}
    return render(request, 'booking.html', context)

# -----------------------------
# --------NOTES----------------
# -----------------------------

def notes(request):
    if not request.user.is_authenticated: return redirect("login")
    form = NoteForm()
    if request.method == 'POST':
        form = NoteForm(request.POST or None)
        if form.is_valid(): form.save()
    object = Note.objects.exclude(parent__isnull=False).first()

    return redirect("note", object.id)

def edit_note(request, id):
    if not request.user.is_authenticated: return redirect("login")
    object = Note.objects.get(id=int(id))
    if request.method == 'POST':
        form = NoteForm(request.POST, instance=object)
        if form.is_valid():
            form.save()
            messages.success(request, "Note saved")
            return redirect("note", object.id)
        else:
            for field in form:
                for error in field.errors:
                    print("Error:", field, error)
    form = NoteForm(instance=object)
    categories = Category.objects.all()
    context = {'object': object, 'categories': categories, 'form': form}
    return render(request, 'note_edit.html', context)

def note(request, id):
    if not request.user.is_authenticated: return redirect("login")
    object = Note.objects.get(id=id)
    if request.method == 'POST':
        form = NoteForm(request.POST or None)
        if form.is_valid():
            new_note = form.save()
        else:
            text = form.cleaned_data['text']
            new_note = Note.objects.create(text=text)
        new_note.parent = object
        new_note.save()
    form_empty = NoteForm()
    form = NoteForm(instance=object)
    object.order_children()
    children = Note.objects.filter(parent=object).order_by('category', 'order')
    # add_order_to_children(object)
    count = len(children)
    categories = Category.objects.all()

    home_pc = socket.gethostname() == "Mum_and_Dads"
    context = {'object': object, 'categories': categories, 'children': children, 'count': count, 'form_empty': form_empty, 'form': form, 'home_pc': home_pc}
    return render(request, 'note.html', context)

def up(request, id):
    return reorder(request, -1, id)

def down(request, id):
    return reorder(request, 1, id)

def reorder(request, dir, id):
    object = Note.objects.filter(id=id).first()
    if object is None:
        pass
    elif dir == 1 and object.order == object.parent.max_child_number():
        pass
    elif dir == -1 and object.order == 1:
        pass
    else:
        if object.order:
            other_object = Note.objects.filter(parent=object.parent, order=object.order + dir).first()
            if other_object:
                other_object.order = object.order
                other_object.save()
            object.order = object.order + dir
        else:
            object.order = object.parent.next_child_number()
        object.save()

    if object and object.parent:
        return redirect("note", str(object.parent.id))
    else:
        return redirect("notes")

def add_order_to_children(object):
    children = Note.objects.filter(parent=object).order_by('category')
    used_numbers = []
    for child in children:
        if child.order is None:
            child.order = object.next_child_number()
            child.save()
        if child.order in used_numbers:
            child.order = object.next_child_number()
            child.save()
        used_numbers.append(child.order)

def delete_note(request, id):
    object = Note.objects.filter(id=id).first()
    if object and object.parent:
        parent = object.parent
    if object:
        object.delete()
    if parent:
        return redirect("note", f"{parent.id}")
    else:
        return redirect("notes")

# -----------------------------
# --------DIARY----------------
# -----------------------------

def diary(request):
    if not request.user.is_authenticated: return redirect("login")
    form = None
    if request.method == 'POST':
        form = DiaryForm(request.POST or None)
        if form.is_valid(): form.save()
    form = DiaryForm()
    objects = Diary.objects.all().order_by("-entry_date")
    count = len(objects)
    context = {'objects': objects, 'title': "Diary", 'count': count, "form": form}
    return render(request, 'diary.html', context)

def diary_delete(request, id):
    object = Diary.objects.filter(id=id).first()
    if object: object.delete()
    return redirect("diary")

# -----------------------------
# --------EVENTS---------------
# -----------------------------
def events(request):
    general = General.objects.get(name="main")
    today = date.today()

    if request.method == 'POST':
        form = EventForm(request.POST or None)
        if form.is_valid(): form.save()

    event_list = []
    for x in range(general.event_days):
        day = today + timedelta(days=x)
        event_list.append([day, [], [], [], []])

    # Events
    events = Event.objects.filter(date__gte=today).order_by('date')
    for event in events:
        for event_day in event_list:
            if event.date == event_day[0]:
                event_day[1].append(event.description)

    # Notes
    notes = Note.objects.filter(note_date__gte=today).order_by('note_date')
    for item in notes:
        for event_day in event_list:
            if item.note_date == event_day[0]:
                event_day[2].append(item)

    # Birthdays
    birthdays = Birthday.objects.all()
    print("Birthdays")
    for item in birthdays:
        print(item, item.date)
        for event_day in event_list:
            if item.date.month == event_day[0].month and item.date.day == event_day[0].day:
                event_day[3].append(item.next_age())

    # Dog Bookings
    bookings = Booking.objects.filter(end_date__gte=today).order_by('start_date')
    for booking in bookings:
        for event_day in event_list:
            if booking.start_date <= event_day[0] <= booking.end_date:
                event_day[4].append(booking.dog.name)

    durations = [7, 30, 60, 90, 120, 150, 180, 360]
    print("\nEvent List")
    print(event_list)

    context = {'event_list': event_list, 'durations': durations, 'general': general, 'title': "Events"}

    return render(request, 'events.html', context)

def event_duration(request, dur):
    general = General.objects.get(name="main")
    general.event_days = dur
    general.save()
    return redirect('events')



# -----------------------------
# --------QUOTES---------------
# -----------------------------

def quotes(request):
    if not request.user.is_authenticated: return redirect("login")
    if request.method == 'POST':
        form = QuoteForm(request.POST or None)
        if form.is_valid(): form.save()
    objects = Quote.objects.all().order_by('-date')
    context = {'quotes': objects}
    return render(request, 'quotes.html', context)

# -----------------------------
# --------BIRTHDAYS-------------
# -----------------------------

def birthdays(request):
    if request.method == 'POST':
        form = BirthdayForm(request.POST or None)
        if form.is_valid(): form.save()
    objects = Birthday.objects.all().order_by(ExtractMonth('date'), ExtractDay('date'))
    count = len(objects)

    context = {'objects': objects, "count": count, "min_days": min_days_to_birthday(), 'title': "Birthdays"}
    return render(request, 'birthdays.html', context)

# -----------------------------
# --------SHOPPING-------------
# -----------------------------

def shopping(request):
    if not request.user.is_authenticated: return redirect("login")
    if request.method == 'POST':
        form = ShoppingForm(request.POST)
        if form.is_valid(): form.save()
    shops = Shop.objects.order_by('order')

    form = ShoppingForm()
    context = {'form': form, 'shops': shops}
    return render(request, 'shopping.html', context)

def shopping_save(request):
    if not request.user.is_authenticated: return redirect("login")
    objects = Shopping.objects.all()
    if request.method == 'POST':
        for object in objects:
            if f"checkbox{object.id}" in request.POST.keys():
                object.buy = True
            else:
                object.buy = False
            object.save()
    return redirect('shopping')

def shopping_clear(request):
    if not request.user.is_authenticated: return redirect("login")
    objects = Shopping.objects.all()
    for object in objects:
        object.buy = False
        object.save()

    return redirect('shopping')
def shopping_delete(request, id):
    object = Note.objects.filter(id=id).first()
    if object and object.shop: parent = object.shop
    if object: object.delete()
    if parent: return redirect("shopping_edit", f"{parent.id}")
    else: return redirect("shopping")

def shopping_edit(request, id):
    shop = Shop.objects.get(id=id)
    shop.order_children()
    context = {'shop': shop}
    return render(request, "shopping_edit.html", context)

def shopping_up(request, id): return shopping_reorder(request, -1, id)

def shopping_down(request, id): return shopping_reorder(request, 1, id)

def shopping_reorder(request, dir, id):
    object = Shopping.objects.filter(id=id).first()
    if object is None:                                                  pass
    elif dir == 1 and object.order == object.shop.max_child_number():   pass
    elif dir == -1 and object.order == 1:                               pass
    else:
        if object.order:
            other_object = Shopping.objects.filter(shop=object.shop, order=object.order + dir).first()
            if other_object:
                other_object.order = object.order
                other_object.save()
            object.order = object.order + dir
        else:
            object.order = object.shop.next_child_number()
        object.save()

    if object and object.shop:
        return redirect("shopping_edit", str(object.shop.id))
    else:
        return redirect("notes")

